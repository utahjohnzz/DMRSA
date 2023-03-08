# -*- coding: utf-8 -*-
"""
Created on Tue Feb 28 14:12:56 2023

@author: utah.johnson
"""




#######Function
def DMRSAmA(lgba, lgdo, lhld,lhto, lsmo, b):
    import numpy as np
    import matplotlib.pyplot as plt
    import os
    import openpyxl
    wb=openpyxl.Workbook()
    ws=wb.active

    path = os.path.join(os.path.expanduser("~"), "Desktop", "Test Exports", "MomentArm.xlsx")



    #rotational matrices
    #z is theta
    #b is the neck angle
    def nRa(z):
        nRa=np.array(([np.cos(np.deg2rad(z)), -np.sin(np.deg2rad(z)), 0], [np.sin(np.deg2rad(z)), np.cos(np.deg2rad(z)), 0], [ 0, 0, 1]))
        return nRa
    def aRb(b):
        bz=180-b
        
        aRb=np.array(([-np.cos(np.deg2rad(bz)), np.sin(np.deg2rad(bz)), 0], [-np.sin(np.deg2rad(bz)), -np.cos(np.deg2rad(bz)), 0], [0, 0, 1]))
        return aRb
    def nRb(z,b):
        nRa=np.array(([np.cos(np.deg2rad(z)), np.sin(np.deg2rad(z)), 0], [np.sin(np.deg2rad(z)), np.cos(np.deg2rad(z)), 0], [ 0, 0, 1]))
        bz=180-b
        aRb=np.array(([-np.cos(np.deg2rad(bz)), np.sin(np.deg2rad(bz)), 0], [-np.sin(np.deg2rad(bz)), -np.cos(np.deg2rad(bz)), 0], [0, 0, 1]))
        nRb=aRb*nRa
        return nRb
    def aRn(z):
        aRn=np.array(([np.cos(np.deg2rad(z)), np.sin(np.deg2rad(z)), 0], [-np.sin(np.deg2rad(z)), np.cos(np.deg2rad(z)), 0], [ 0, 0, 1]))
        return aRn
    def alpha(b):
        b=b-135
        alpha=np.array(([np.cos(np.deg2rad(b)), -np.sin(np.deg2rad(b)), 0], [np.sin(np.deg2rad(b)), np.cos(np.deg2rad(b)), 0], [ 0, 0, 1]))
        return alpha



    ####################
    #Glenohumeral Joint Vectors
    #all vectors will be WRT to N coordinate system using rotational matrices
    #Pgba is static vector of glenoid offset
    
    Pgba=np.array(([0, lgba, 0])) #no need for rotation
    
    lhdp=10 #length from greater tuberosity to proximal deltoid insertion
    lhdi=42.3/2 #diameter of humeral bone
    #user granted step data based on computational speed
    step=91
    #creating abduction angle range
    abrange=90
    abang=np.linspace(0,abrange,step)
    #we need to iterate over the range to find the indexwise moment 
    c1=0
    Pgdo=np.zeros((1,1))
    indexd=np.zeros((1,))
    indexp=np.zeros((1,))
    indexpsh=np.zeros((1,))
    indec=0
    for z in abang:
        Pgdo=np.array(nRa(z)*[0, -lgdo/2, 0]) #index vector for glenoid diameter in N coordinates
        Pgdo=Pgdo[:,1]
        Phld=nRa(z)*[0, -lhld, 0] #index vector for humeral liner depth in N coordinates
        Phld=Phld[:,1]
        Phto=nRa(z)*[0, -lhto, 0] #index vector for humeral tray offset in N coordinates
        Phto=Phto[:,1]
        Psmo=nRb(z,b)*[-lsmo, 0, 0] #index vector for medial humeral offset in N coordinates
        Psmo=Psmo[:,0]
        Psh=Pgdo+Phld+Phto+Psmo
        mash=Psh[0,]
        indexpsh=np.append(indexpsh,mash)
        zz=abs(z-60)
        
        #### humeral component
        #need vectors of both the humeral diameter and the humeral length
        Psi=nRa(z)*alpha(b)*[lhdi,lhdp,0]
        maxx=Psi[0,0]+Psi[0,1]
        momentarm=(maxx+mash)
        if zz<1:
            masix=momentarm
            angle=z
        indexd=np.append(indexd, momentarm)   
        if momentarm>indec:
            indec=momentarm
            enc=z
    indexd=np.delete(indexd,(0,0))
    indexpsh=np.delete(indexpsh,(0,0))
    plt.plot(abang,indexd,marker='o')
    #plt.plot(abang,indexpsh)
    plt.title('Moment Arm of Medial Deltoid During Abduction')
    plt.xlabel('Abduction Angle (Degree)')
    plt.ylabel('Moment Arm (mm)')

    print('Maximum moment arm of',round(indec,2),'mm at',round(enc,2),'degrees')
    print('Moment arm at minimum is',indexd[0],'and the moment arm at 60 degrees is',masix)

    ws['A1']='Glenoid Diameter'
    ws['A2']=lgdo
    ws['B1']='Humeral Liner Depth'
    ws['B2']=lhld
    ws['C1']='Humeral Tray Offset'
    ws['C2']=lhto
    ws['D1']='Humeral Medial Offset'
    ws['D2']=lsmo
    ws['E1']='Humeral Neck Shaft Angle'
    ws['E2']=b
    ws['F1']='Maximum Moment Arm'
    ws['G1']='Maximum Moment Arm Location'
    ws['H1']='Minimum Moment Arm'
    ws['I1']='Moment Arm at 60 Degrees'
    wb.save(path)


lgba=[38,42,46]
lgdo=[0, 3.5, 2.5, 4, 2.3, 6, 9.5,6.3, -2, 5.8, 8.5]
lhld=[18.5,18.6,18.8,18.9]
lhto=[0, 2.5, 5, 7.5, 10,12.5, 15, 17.5]
lsmo=[7.5, 8.5, 9.5]
b=[135,145,155]


import random
import itertools

# define the trial variables


# combine the variables into a list of tuples
all_trials = list(itertools.product(lgba, lgdo, lhld, lhto,lsmo,b))

# randomize the order of the trials
random.shuffle(all_trials)
# iterate over the randomized trials
for trial in all_trials:
    DMRSAmA(trial[0],trial[1],trial[2],trial[3],trial[4],trial[5])
    

